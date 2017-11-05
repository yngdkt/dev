# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font

class ModelConfig(object):
    def __init__(self):
        pass
    
    def test_func(self, val):
        return val*2
    

    
    
    
wb = Workbook()
ws = wb.active
ws.sheet_view.showGridLines = False


# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()



# Save the file
wb.save("C:\\temp\\sample_hoge.xlsx")