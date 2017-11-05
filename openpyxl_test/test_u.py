# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

class ModelConfig(object):
    def __init__(self):
        pass
    
    def test_func(self, val):
        return val*2
    

    
wb = load_workbook(filename='C:\\temp\\blank.xlsx')    
    
#wb = Workbook()
ws = wb.active

#ft1 = Font(name='Arial',size=9)

# Data can be assigned directly to cells
#ws['A1'] = "ぴよぴよ"

ws.column_dimensions["B"].width = 2.0
ws.column_dimensions["C"].width = 2.0
ws.column_dimensions["D"].width = 2.0
ws.column_dimensions["E"].width = 2.0
ws.column_dimensions["F"].width = 2.0
ws.column_dimensions["G"].width = 15.0

ts_header_list = ["16/03", "17/03", "18/03", "19/03", "20/03", "21/03"]

ws["H2"] = ts_header_list[0]
ws["I2"] = ts_header_list[1]
ws["J2"] = ts_header_list[2]
ws["K2"] = ts_header_list[3]
ws["L2"] = ts_header_list[4]
ws["M2"] = ts_header_list[5]

ft_ts_header = Font(underline = 'singleAccounting')

al_right = Alignment(horizontal = "right")
ws["H2"].font = ft_ts_header
ws["I2"].font = ft_ts_header
ws["J2"].font = ft_ts_header
ws["K2"].font = ft_ts_header
ws["L2"].font = ft_ts_header
ws["M2"].font = ft_ts_header

ws["H2"].alignment = al_right
ws["I2"].alignment = al_right
ws["J2"].alignment = al_right
ws["K2"].alignment = al_right
ws["L2"].alignment = al_right
ws["M2"].alignment = al_right

ws["H3"] = ""

sales = (1028.38500,1178.29000,1199.31100,1295.25588,1398.87635,1496.797695)
ws.append(("",)* 7 + sales)

ws["H4"].number_format = "#,##0_);(#,##0)"
ws["I4"].number_format = "#,##0_);(#,##0)"
ws["J4"].number_format = "#,##0_);(#,##0)"
ws["K4"].number_format = "#,##0_);(#,##0)"
ws["L4"].number_format = "#,##0_);(#,##0)"
ws["M4"].number_format = "#,##0_);(#,##0)"

# Rows can also be appended
#ws.append([1, 2, 3])
#ws['A2'].font = ft1

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()



# Save the file
wb.save("C:\\temp\\sample_hoge.xlsx")